"""
run_simulation.py — LitmusAI conversation simulator.

For every persona × company pair:
  1. Generates a persona system prompt via Gemma4 (the "human tester" LLM).
  2. Builds a chatbot system prompt from the company profile.
  3. Runs an N-turn conversation: persona LLM ↔ chatbot LLM.
  4. Saves the full conversation transcript to simulation_results.xlsx.

Usage (from the personas/ directory):
    python run_simulation.py [--turns N]

Or from the project root:
    python personas/run_simulation.py [--turns N]
"""

import argparse
import re
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests

# ---------------------------------------------------------------------------
# Path setup — works whether run from personas/ or the project root
# ---------------------------------------------------------------------------

PERSONAS_DIR = Path(__file__).parent
sys.path.insert(0, str(PERSONAS_DIR))

TESTER_PROFILES_DIR = PERSONAS_DIR / "tester_profiles"
COMPANIES_DIR = PERSONAS_DIR / "companies"
COMPANIES_OVERVIEW = PERSONAS_DIR / "companies_overview.md"
RESULTS_FILE = PERSONAS_DIR / "simulation_results.xlsx"

# ---------------------------------------------------------------------------
# Ollama client
# ---------------------------------------------------------------------------

OLLAMA_URL = "http://localhost:11434/api/chat"
MODEL = "gemma4:latest"


def query_gemma_chat(system_msg: str, messages: list[dict]) -> str:
    """Multi-turn call — messages is the full conversation history so far."""
    payload = {
        "model": MODEL,
        "messages": [{"role": "system", "content": system_msg}] + messages,
        "stream": False,
    }
    response = requests.post(OLLAMA_URL, json=payload, timeout=300)
    response.raise_for_status()
    return response.json()["message"]["content"]


# ---------------------------------------------------------------------------
# Markdown helpers
# ---------------------------------------------------------------------------


def _extract_section(md: str, *header_variants: str) -> str:
    """
    Return the body of the first ## section whose heading matches any of the
    provided variants (case-insensitive, partial match allowed).
    Stops at the next ## heading.
    """
    lines = md.splitlines()
    body: list[str] = []
    capturing = False

    for line in lines:
        if line.startswith("## "):
            heading = line[3:].strip().lower()
            if capturing:
                break
            if any(v.lower() in heading for v in header_variants):
                capturing = True
            continue
        if capturing:
            body.append(line)

    return "\n".join(body).strip()


def _first_sentence(text: str) -> str:
    """Return the first sentence of a block of text."""
    text = re.sub(r"\s+", " ", text.replace("\n", " ")).strip()
    match = re.match(r"([^.!?]+[.!?])", text)
    return match.group(1).strip() if match else text[:120]


def _extract_inline(md: str, bold_label: str) -> str:
    """Extract the value after a **Label:** pattern anywhere in the file."""
    pattern = rf"\*\*{re.escape(bold_label)}:\*\*\s*(.+)"
    match = re.search(pattern, md)
    return match.group(1).strip() if match else ""


# ---------------------------------------------------------------------------
# Parsers
# ---------------------------------------------------------------------------


def parse_persona(path: Path) -> dict:
    md = path.read_text(encoding="utf-8")

    name_match = re.match(r"#\s+Persona:\s+(.+)", md)
    name = name_match.group(1).strip() if name_match else path.stem

    identity_block = _extract_section(md, "identity")
    behavioural = _extract_section(md, "behavioural profile", "behavioral profile")
    summary = _extract_section(md, "summary")
    tone_raw = _extract_section(md, "tone")

    brief_identity = _first_sentence(summary) if summary else _first_sentence(identity_block)
    brief_identity = re.sub(r"\*\*", "", brief_identity)

    full_identity = "\n\n".join(filter(None, [identity_block, behavioural]))

    raw_openers = _extract_section(md, "example openers")
    opener_lines = [
        line.strip().lstrip("- ").strip('"')
        for line in raw_openers.splitlines()
        if line.strip().startswith("- ")
    ]
    example_openers = "\n".join(f'- "{o}"' for o in opener_lines if o)

    return {
        "name": name,
        "brief_identity": brief_identity,
        "identity_block": full_identity or identity_block,
        "tone": re.sub(r"\*\*", "", tone_raw).strip(),
        "example_openers": example_openers,
        "interaction_rules": _extract_section(md, "interaction rules"),
        "failure_patterns": _extract_section(md, "failure patterns"),
        "role_anchor": _extract_section(md, "role anchor"),
    }


def _parse_industry_map() -> dict[str, str]:
    """Read companies_overview.md and build {company_name: industry} mapping."""
    industry_map: dict[str, str] = {}
    if not COMPANIES_OVERVIEW.exists():
        return industry_map
    md = COMPANIES_OVERVIEW.read_text(encoding="utf-8")
    for block in re.split(r"^---", md, flags=re.MULTILINE):
        name_match = re.search(r"##\s+C\d+\s+[—–-]+\s+(.+)", block)
        industry_match = re.search(r"\*\*Industry:\*\*\s*(.+)", block)
        if name_match and industry_match:
            industry_map[name_match.group(1).strip()] = industry_match.group(1).strip()
    return industry_map


_INDUSTRY_MAP: dict[str, str] = {}


def parse_company(path: Path) -> dict:
    global _INDUSTRY_MAP
    if not _INDUSTRY_MAP:
        _INDUSTRY_MAP = _parse_industry_map()

    md = path.read_text(encoding="utf-8")

    name_match = re.match(r"#\s+Company Profile:\s+(.+)", md)
    company_name = name_match.group(1).strip() if name_match else path.stem

    agent_section = _extract_section(md, "2.", "the agent")
    bot_name = _extract_inline(agent_section, "Name")
    if not bot_name:
        bot_name = f"{company_name} Assistant"

    industry = _INDUSTRY_MAP.get(company_name, "")
    capabilities = _extract_section(md, "3.", "core capabilities")
    test_scenarios = _extract_section(md, "4.", "test case parameters")

    system_rules = _extract_section(md, "5.", "system instructions")

    return {
        "bot_name": bot_name,
        "company_name": company_name,
        "industry": industry,
        "capabilities_summary": capabilities,
        "test_scenarios": test_scenarios,
        "system_rules": system_rules,
    }


# ---------------------------------------------------------------------------
# Persona prompt builder (directly from parsed file — no LLM generation)
# ---------------------------------------------------------------------------


def build_persona_prompt(persona: dict) -> str:
    """Construct a persona system prompt directly from the tester profile file."""
    parts: list[str] = []

    if persona.get("role_anchor"):
        parts.append(persona["role_anchor"].strip())

    if persona.get("identity_block"):
        parts += ["", "## Who You Are", persona["identity_block"].strip()]

    if persona.get("tone"):
        parts += ["", "## Tone", persona["tone"].strip()]

    if persona.get("interaction_rules"):
        parts += ["", "## Interaction Rules", persona["interaction_rules"].strip()]

    if persona.get("failure_patterns"):
        parts += ["", "## Failure Patterns (apply these naturally)", persona["failure_patterns"].strip()]

    if persona.get("example_openers"):
        parts += ["", "## Example Openers (use one to start, then improvise)", persona["example_openers"].strip()]

    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Chatbot prompt builder
# ---------------------------------------------------------------------------


def build_chatbot_prompt(company: dict) -> str:
    """Build a system prompt for the chatbot LLM from the company profile."""
    parts = [
        f"You are {company['bot_name']}, the official customer service chatbot for {company['company_name']}.",
        "",
    ]
    if company.get("capabilities_summary"):
        parts += ["## Your Capabilities", company["capabilities_summary"].strip(), ""]
    if company.get("system_rules"):
        parts += ["## Rules You Must Follow", company["system_rules"].strip(), ""]
    parts += [
        "Respond helpfully and concisely. Stay fully in character as the chatbot at all times.",
        "Never break character or acknowledge that you are an LLM or part of a simulation.",
        "You will NEVER use internal formatting syntax like markdown bold or headers in your replies — "
        "respond in plain conversational text only.",
    ]
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Conversation runner
# ---------------------------------------------------------------------------


def run_conversation(
    persona_prompt: str,
    chatbot_prompt: str,
    persona_name: str,
    bot_name: str,
    n_turns: int = 8,
) -> list[dict]:
    """Run a full multi-turn conversation between the persona and the chatbot.

    Each turn: the persona LLM generates a message, then the chatbot LLM replies.
    The history is maintained separately for each side so each model sees itself
    as "assistant" and the other as "user".

    Returns a list of {"persona": str, "bot": str} dicts, one per turn.
    """
    transcript: list[dict] = []

    for turn_num in range(1, n_turns + 1):
        print(f"\n  --- Turn {turn_num}/{n_turns} ---")

        # Persona's view: its own past messages are "assistant", bot replies are "user"
        persona_messages: list[dict] = []
        for t in transcript:
            persona_messages.append({"role": "assistant", "content": t["persona"]})
            persona_messages.append({"role": "user", "content": t["bot"]})

        print(f"  {persona_name}: ", end="", flush=True)
        persona_msg = query_gemma_chat(persona_prompt, persona_messages)
        print(persona_msg.strip())

        # Chatbot's view: persona messages are "user", its own past replies are "assistant"
        chatbot_messages: list[dict] = []
        for t in transcript:
            chatbot_messages.append({"role": "user", "content": t["persona"]})
            chatbot_messages.append({"role": "assistant", "content": t["bot"]})
        chatbot_messages.append({"role": "user", "content": persona_msg})

        print(f"  {bot_name}: ", end="", flush=True)
        chatbot_msg = query_gemma_chat(chatbot_prompt, chatbot_messages)
        print(chatbot_msg.strip())

        transcript.append({"persona": persona_msg, "bot": chatbot_msg})

    return transcript


def format_transcript(transcript: list[dict]) -> str:
    """Format a conversation transcript as a readable string for Excel."""
    lines: list[str] = []
    for i, turn in enumerate(transcript, start=1):
        lines.append(f"[Turn {i}]")
        lines.append(f"PERSONA: {turn['persona'].strip()}")
        lines.append(f"BOT:     {turn['bot'].strip()}")
        lines.append("")
    return "\n".join(lines).strip()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

HEADERS = ["#", "Persona", "Company", "Turns", "Elapsed (s)", "Conversation Transcript", "Error"]
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ERROR_FILL = PatternFill("solid", fgColor="FFE0E0")


def _init_excel() -> None:
    """Create the results file with headers and column widths."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulation Transcripts"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.column_dimensions["A"].width = 5    # #
    ws.column_dimensions["B"].width = 16   # Persona
    ws.column_dimensions["C"].width = 18   # Company
    ws.column_dimensions["D"].width = 8    # Turns
    ws.column_dimensions["E"].width = 12   # Elapsed
    ws.column_dimensions["F"].width = 120  # Conversation Transcript
    ws.column_dimensions["G"].width = 40   # Error

    wb.save(RESULTS_FILE)


def _append_row(result: dict, row_num: int) -> None:
    """Append a single result row to the existing Excel file."""
    wb = openpyxl.load_workbook(RESULTS_FILE)
    ws = wb.active
    is_error = bool(result.get("error"))
    row = [
        row_num,
        result.get("persona", ""),
        result.get("company", ""),
        result.get("turns", ""),
        result.get("elapsed_s", ""),
        result.get("transcript", "") if not is_error else "",
        result.get("error", ""),
    ]
    excel_row = row_num + 1  # +1 for header row
    for col, value in enumerate(row, start=1):
        cell = ws.cell(row=excel_row, column=col, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if is_error:
            cell.fill = ERROR_FILL
    wb.save(RESULTS_FILE)


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------


def run_all(n_turns: int = 8) -> None:
    personas = sorted(TESTER_PROFILES_DIR.glob("p*.md"))
    companies = sorted(COMPANIES_DIR.glob("c*.md"))

    if not personas:
        print(f"No persona files found in {TESTER_PROFILES_DIR}")
        sys.exit(1)
    if not companies:
        print(f"No company files found in {COMPANIES_DIR}")
        sys.exit(1)

    persona_dicts = [parse_persona(p) for p in personas]
    company_dicts = [parse_company(c) for c in companies]

    total = len(persona_dicts) * len(company_dicts)
    print(
        f"Running simulations: {len(persona_dicts)} personas × {len(company_dicts)} companies"
        f" = {total} conversations ({n_turns} turns each)\n"
    )

    _init_excel()
    count = 0

    for persona in persona_dicts:
        for company in company_dicts:
            count += 1
            label = f"[{count:02d}/{total}] {persona['name']} × {company['company_name']}"
            print(f"{label}")

            try:
                t0 = time.time()

                # Step 1: Build prompts directly from profile files (no LLM generation)
                persona_prompt = build_persona_prompt(persona)
                chatbot_prompt = build_chatbot_prompt(company)

                # Step 2: Run the conversation
                t1 = time.time()
                transcript = run_conversation(
                    persona_prompt, chatbot_prompt,
                    persona_name=persona["name"],
                    bot_name=company["bot_name"],
                    n_turns=n_turns,
                )
                elapsed = round(time.time() - t0, 1)
                print(f"\n  Finished in {elapsed}s")

                result = {
                    "persona": persona["name"],
                    "company": company["company_name"],
                    "turns": n_turns,
                    "elapsed_s": elapsed,
                    "transcript": format_transcript(transcript),
                }

            except requests.exceptions.ConnectionError:
                print("FAILED — Ollama not reachable at http://localhost:11434")
                print("  Make sure Ollama is running: ollama serve")
                sys.exit(1)
            except Exception as exc:
                print(f"  ERROR — {exc}")
                result = {
                    "persona": persona["name"],
                    "company": company["company_name"],
                    "error": str(exc),
                }

            _append_row(result, count)

    print(f"\n{'='*60}")
    print(f"Done. {count} conversations saved to {RESULTS_FILE.name}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="LitmusAI conversation simulator")
    parser.add_argument(
        "--turns", type=int, default=8,
        help="Number of conversation turns per simulation (default: 8)",
    )
    args = parser.parse_args()
    run_all(n_turns=args.turns)
